# Scene Task Data Visualizer
# Created 6/3/15, Sam Weiller
# Direct questions to sam.weiller@gmail.com
# v. 0.0 (Versions maintained by GitHub)

# ***** MODULE INITIALIZATION ***** #
import PIL
from PIL import Image, ImageDraw
import openpyxl as pxl
import numpy as np
import os
import progressbar
from colorama import Fore, Back, Style, init
import datetime
import sys

init(autoreset=True)

# ***** CUSTOM FUNCTIONS ***** #
def logInformation(textToPrint):
	prefix = 'LOG ' + datetime.time.isoformat(datetime.datetime.now().time()) + ':    '
	print(Fore.BLACK + Back.GREEN + Style.BRIGHT + prefix + textToPrint)

logInformation('Script is active.')
logInformation('This is a message.')

def logWarning(textToPrint):
	prefix = 'WRN ' + datetime.time.isoformat(datetime.datetime.now().time()) + ':    '
	print(Fore.BLACK + Back.YELLOW + Style.BRIGHT + prefix + textToPrint)

logWarning('This is a warning.')


# ***** CONTROL PANEL ***** #
fixationReportName = '/Users/samweiller/Dropbox-Dilks/Dropbox/SCTbehavioral/trialImages/freeviewFixReport603forAnalysis.xlsx'
# !! NOTE: The fix report must be an xlsx file. openpyxl cannot read xls files.
imageDirectory = '/Users/samweiller/Dropbox/SCBimages'
savedImagePath = '/Users/samweiller/Desktop/SCBanalysisImages'

navCoords = {}
catCoords = {}
colCoords = {}

dotSize = 20
dotFactor = round(dotSize/2)
transparencyFactor = 50
navColor = (255, 255, 0, transparencyFactor)
catColor = (0, 255, 0, transparencyFactor)
colColor = (0, 0, 255, transparencyFactor)

logInformation('Initialization is complete. Beginning the main script.')

# ***** IMPORT DATA ***** #
# Import the fixation report
logInformation('Loading in the fixation report.')
fixReportExtension = os.path.splitext(fixationReportName)
if fixReportExtension[1] == '.xlsx':
	theWorkbook = pxl.load_workbook(fixationReportName)
	fixationData = theWorkbook.active
	logInformation('Fixation report is loaded in.')
else:
	raise Exception('Incorrect file type for fixation report. Must be xlsx.')

# Check and verify the report
validNameOrder = ['RECORDING_SESSION_LABEL', 'TRIAL_INDEX', 'CURRENT_FIX_INDEX', 'CURRENT_FIX_X', 'CURRENT_FIX_Y', 'TASK', 'DIRECTION', 'CATEGORY', 'COLOR', 'IMG_NAME', 'CURRENT_FIX_PUPIL'];
actualNameOrder = [];
for columnNumber in range(1,12):
	actualNameOrder.append(fixationData.cell(row = 1, column = columnNumber).value)

if actualNameOrder == validNameOrder:
	logInformation('Column name order is valid. Moving on.')
else:
	raise Exception('One or more column names are invalid or in the wrong order. Please recreate your fixation report.')

# ***** IMPORT IMAGES ***** #
imageNames = [os.path.join(imageDirectory, f) for f in os.listdir(imageDirectory)]

# ***** CREATE COORDINATE MATRICES ***** #
for image in imageNames:
	if image.endswith('bmp'):
		navCoords[os.path.splitext(os.path.split(image)[1])[0]] = []
		catCoords[os.path.splitext(os.path.split(image)[1])[0]] = []
		colCoords[os.path.splitext(os.path.split(image)[1])[0]] = []
	else:
		logWarning('File ' + os.path.split(image)[1] + ' is not a bmp.')


pbar = progressbar.ProgressBar().start()
logInformation('Processing fixations.')

for rowNumber in pbar(fixationData.rows): #create tuple lists
	if os.path.splitext(str(rowNumber[9].value))[0] == None:
		dummy = 0
		# logWarning('Row has a value of Null.')
	elif os.path.splitext(str(rowNumber[9].value))[0] == '.':
		dummy = 0
		# logWarning('Row belongs to a fixation period.')
	else:
		if rowNumber[5].value == 'NAVIGATION':
			navCoords[os.path.splitext(str(rowNumber[9].value))[0]].append((rowNumber[3].value, rowNumber[4].value))
		elif rowNumber[5].value == 'CATEGORIZATION':
			catCoords[os.path.splitext(str(rowNumber[9].value))[0]].append((rowNumber[3].value, rowNumber[4].value))
		elif rowNumber[5].value == 'COLOR':
			colCoords[os.path.splitext(str(rowNumber[9].value))[0]].append((rowNumber[3].value, rowNumber[4].value))

logInformation('Applying fixations to images.')
for image in imageNames:
	logInformation('Applying fixations to image' + os.path.splitext(os.path.split(image)[1])[0] + '.')

	theOpenImage = Image.open(image)
	compositeImage = theOpenImage
	bigCompositeImage = theOpenImage
	medCompositeImage = theOpenImage

	logInformation('Processing Color.')
	pbar2 = progressbar.ProgressBar().start()
	for eachPoint in pbar2(colCoords[os.path.splitext(os.path.split(image)[1])[0]]):
		theMask = Image.new('RGBA', (1280, 1023), 0) # images are 1280x1024
		draw = ImageDraw.Draw(theMask)
		# draw.point(eachPoint)
		draw.ellipse((round(eachPoint[0])-dotFactor, round(eachPoint[1])-dotFactor, round(eachPoint[0])+dotFactor, round(eachPoint[1])+dotFactor), fill=colColor)
		compositeImage = Image.composite(theMask, compositeImage, theMask)
		bigCompositeImage = Image.composite(theMask, bigCompositeImage, theMask)
		del theMask
		del draw
	compositeImage.save(savedImagePath + '/COL_' + os.path.split(image)[1], 'PNG')
	
	logInformation('Processing Categorization.')
	pbar3 = progressbar.ProgressBar().start()
	compositeImage = theOpenImage
	for eachPoint in pbar3(catCoords[os.path.splitext(os.path.split(image)[1])[0]]):
		theMask = Image.new('RGBA', (1280, 1023), 0) # images are 1280x1024
		draw = ImageDraw.Draw(theMask)
		# draw.point(eachPoint)
		draw.ellipse((round(eachPoint[0])-dotFactor, round(eachPoint[1])-dotFactor, round(eachPoint[0])+dotFactor, round(eachPoint[1])+dotFactor), fill=catColor)
		compositeImage = Image.composite(theMask, compositeImage, theMask)
		bigCompositeImage = Image.composite(theMask, bigCompositeImage, theMask)
		medCompositeImage = Image.composite(theMask, medCompositeImage, theMask)
		del theMask
		del draw
	compositeImage.save(savedImagePath + '/CAT_' + os.path.split(image)[1], 'PNG')

	logInformation('Processing Navigation.')
	pbar4 = progressbar.ProgressBar().start()
	compositeImage = theOpenImage
	for eachPoint in pbar4(navCoords[os.path.splitext(os.path.split(image)[1])[0]]):
		theMask = Image.new('RGBA', (1280, 1023), 0) # images are 1280x1024
		draw = ImageDraw.Draw(theMask)
		# draw.point(eachPoint)
		draw.ellipse((round(eachPoint[0])-dotFactor, round(eachPoint[1])-dotFactor, round(eachPoint[0])+dotFactor, round(eachPoint[1])+dotFactor), fill=navColor)
		compositeImage = Image.composite(theMask, compositeImage, theMask)
		bigCompositeImage = Image.composite(theMask, bigCompositeImage, theMask)
		medCompositeImage = Image.composite(theMask, medCompositeImage, theMask)
		del theMask
		del draw
	compositeImage.save(savedImagePath + '/NAV_' + os.path.split(image)[1], 'PNG')

	logInformation('Saving Composite Images.')
	bigCompositeImage.save(savedImagePath + '/ALL_' + os.path.split(image)[1], 'PNG')
	medCompositeImage.save(savedImagePath + '/NAVCAT' + os.path.split(image)[1], 'PNG')

	


	# compositeImage = Image.composite(theMask, theOpenImage, theMask) #THIS WORKS!	
		
#fixationData.cell(row = rowNumber, column = 6)

# ***** APPLY MATRICES TO IMAGES ***** #

# ***** SAVE THAT SHIT OUT ***** #













